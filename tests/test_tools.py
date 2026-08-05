import asyncio
import base64
import io
import json
from pathlib import Path

from PIL import Image
from mcp.server.fastmcp import FastMCP

from config import Settings
from tools import register_tools


def _data_url(color=(10, 20, 30), width=80, height=60):
    buf = io.BytesIO()
    Image.new("RGB", (width, height), color).save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


def _make_mcp():
    settings = Settings(mock=True, auto_tile=False, max_pixels=1_000_000)
    mcp = FastMCP("vision-mcp-test")
    register_tools(mcp, settings)
    return mcp


def test_all_tools_registered():
    mcp = _make_mcp()

    async def run():
        tools = await mcp.list_tools()
        return {t.name for t in tools}

    names = asyncio.run(run())
    assert names == {
        "analyze_image",
        "image_info",
        "ocr_image",
        "table_from_image",
        "analyze_ui",
        "analyze_document_slide",
        "describe_chart",
        "compare_images",
        "tile_image",
        "parse_document",
        "parse_document_status",
        "analyze_any",
        "scan_folder",
    }


def _call(mcp, name, arguments):
    async def run():
        result = await mcp.call_tool(name, arguments)
        if isinstance(result, tuple):
            result = result[0]
        if isinstance(result, dict):
            return json.dumps(result, ensure_ascii=False)
        return "".join(c.text or "" for c in result)

    return asyncio.run(run())


def test_analyze_image_mock():
    mcp = _make_mcp()
    text = _call(mcp, "analyze_image", {"image": _data_url(), "prompt": "是什么"})
    assert "[mock]" in text


def test_image_info_mock():
    mcp = _make_mcp()
    text = _call(mcp, "image_info", {"image": _data_url()})
    assert '"width": 80' in text
    assert '"height": 60' in text


def test_ocr_and_compare_mock():
    mcp = _make_mcp()
    ocr = _call(mcp, "ocr_image", {"image": _data_url()})
    assert "[mock]" in ocr
    cmp = _call(mcp, "compare_images", {"image_a": _data_url(), "image_b": _data_url()})
    assert "[mock]" in cmp


def test_bad_source_returns_friendly_error():
    mcp = _make_mcp()
    text = _call(mcp, "analyze_image", {"image": "C:/no/such.png"})
    assert text.startswith("错误：")
    assert "文件不存在" in text


def test_auto_tile_splits_large_image():
    settings = Settings(
        mock=True, auto_tile=True, tile_threshold=100, tile_size=64, tile_overlap=0
    )
    mcp = FastMCP("vision-mcp-test")
    register_tools(mcp, settings)
    text = _call(mcp, "analyze_image", {"image": _data_url(width=200, height=100)})
    assert "【块 1/8】" in text
    assert "【块 4/8】" in text


def test_force_tile_tool():
    settings = Settings(
        mock=True, auto_tile=False, tile_overlap=0, max_pixels=1_000_000
    )
    mcp = FastMCP("vision-mcp-test")
    register_tools(mcp, settings)
    text = _call(
        mcp,
        "tile_image",
        {"image": _data_url(width=150, height=150), "tile_size": 100},
    )
    assert "【块 1/4】" in text
    assert "【块 4/4】" in text


def test_auto_tile_off_sends_single():
    settings = Settings(mock=True, auto_tile=False, max_pixels=1_000_000)
    mcp = FastMCP("vision-mcp-test")
    register_tools(mcp, settings)
    text = _call(mcp, "analyze_image", {"image": _data_url(width=200, height=100)})
    assert "已收到 1 张图片" in text


def test_table_export_xlsx(tmp_path):
    settings = Settings(
        mock=False,
        auto_tile=False,
        max_pixels=1_000_000,
        output_dir=tmp_path,
    )

    class TableMock:
        async def complete(self, prompt, images):
            return "| Name | Score |\n|---|---|\n| Alice | 92 |\n| Bob | 85 |"

    import tools as tools_module

    original = tools_module.get_provider
    tools_module.get_provider = lambda s: TableMock()
    try:
        mcp = FastMCP("vision-mcp-test")
        tools_module.register_tools(mcp, settings)
        text = _call(
            mcp,
            "table_from_image",
            {"image": _data_url(), "export_xlsx": True, "out_path": "out.xlsx"},
        )
    finally:
        tools_module.get_provider = original

    assert "表格识别结果" in text
    assert "已导出 Excel" in text
    out = tmp_path / "out.xlsx"
    assert out.is_file()

    from openpyxl import load_workbook

    ws = load_workbook(out).active
    assert ws["A1"].value == "Name"
    assert ws["B2"].value == "92"
    assert ws["B3"].value == "85"


class FakeAutoProvider:
    def __init__(self, detections=None, analyses=None):
        self.detections = list(detections or [])
        self.analyses = list(analyses or [])
        self.detection_calls = 0

    async def complete(self, prompt, images):
        if prompt.startswith("请判断这张图片"):
            self.detection_calls += 1
            return self.detections.pop(0) if self.detections else "general"
        return self.analyses.pop(0) if self.analyses else "[mock] analysis"


def _auto_mcp(tmp_path, fake, monkeypatch):
    import tools as tools_module

    settings = Settings(
        auto_tile=False,
        max_pixels=1_000_000,
        cache_dir=tmp_path / "cache",
    )
    monkeypatch.setattr(tools_module, "get_provider", lambda s: fake)
    mcp = FastMCP("vision-mcp-test")
    tools_module.register_tools(mcp, settings)
    return mcp


def test_analyze_any_detects_and_routes(monkeypatch, tmp_path):
    import tools as tools_module

    fake = FakeAutoProvider(detections=["ui"], analyses=["UI 分析结果"])
    monkeypatch.setattr(tools_module, "get_provider", lambda s: fake)
    mcp = _auto_mcp(tmp_path, fake, monkeypatch)
    text = _call(mcp, "analyze_any", {"source": _data_url()})
    assert "图片类型：ui" in text
    assert "UI 分析结果" in text
    assert fake.detection_calls == 1


def test_analyze_any_hint_skips_detection(monkeypatch, tmp_path):
    import tools as tools_module

    fake = FakeAutoProvider(detections=["ui"], analyses=["表格结果"])
    monkeypatch.setattr(tools_module, "get_provider", lambda s: fake)
    mcp = _auto_mcp(tmp_path, fake, monkeypatch)
    text = _call(mcp, "analyze_any", {"source": _data_url(), "hint": "table"})
    assert "图片类型：table" in text
    assert "表格结果" in text
    assert fake.detection_calls == 0


def test_analyze_any_falls_back_to_general(monkeypatch, tmp_path):
    import tools as tools_module

    fake = FakeAutoProvider(detections=["看不懂"], analyses=["通用结果"])
    monkeypatch.setattr(tools_module, "get_provider", lambda s: fake)
    mcp = _auto_mcp(tmp_path, fake, monkeypatch)
    text = _call(mcp, "analyze_any", {"source": _data_url()})
    assert "图片类型：general" in text
    assert "通用结果" in text


def test_analyze_any_cache_hit(monkeypatch, tmp_path):
    import tools as tools_module

    fake = FakeAutoProvider(detections=["ui"], analyses=["第一次结果"])
    monkeypatch.setattr(tools_module, "get_provider", lambda s: fake)
    mcp = _auto_mcp(tmp_path, fake, monkeypatch)
    first = _call(mcp, "analyze_any", {"source": _data_url(), "hint": "ui"})
    second = _call(mcp, "analyze_any", {"source": _data_url(), "hint": "ui"})
    assert "第一次结果" in first
    assert "【缓存】" in second
    assert "第一次结果" in second


def test_scan_folder_registered():
    mcp = _make_mcp()

    async def run():
        tools = await mcp.list_tools()
        return {t.name for t in tools}

    names = asyncio.run(run())
    assert "scan_folder" in names


def _make_png(path, color=(10, 20, 30), width=40, height=30):
    Image.new("RGB", (width, height), color).save(path, format="PNG")


def test_scan_folder_analyzes_new_images(monkeypatch, tmp_path):
    import tools as tools_module

    folder = tmp_path / "shots"
    folder.mkdir()
    _make_png(folder / "a.png", (1, 2, 3))
    _make_png(folder / "b.png", (4, 5, 6))
    fake = FakeAutoProvider(detections=["ui", "table"], analyses=["A 分析", "B 分析"])
    monkeypatch.setattr(tools_module, "get_provider", lambda s: fake)
    settings = Settings(
        auto_tile=False, max_pixels=1_000_000, cache_dir=tmp_path / "cache"
    )
    mcp = FastMCP("vision-mcp-test")
    tools_module.register_tools(mcp, settings)

    text = _call(mcp, "scan_folder", {"folder": str(folder)})
    assert "a.png" in text
    assert "b.png" in text
    assert "A 分析" in text
    assert "B 分析" in text

    text2 = _call(mcp, "scan_folder", {"folder": str(folder)})
    assert "没有新图片" in text2
    assert (tmp_path / "cache" / "scan_cursor.json").is_file()


def test_scan_folder_since_filter(monkeypatch, tmp_path):
    import os

    import tools as tools_module

    folder = tmp_path / "shots"
    folder.mkdir()
    f1 = folder / "old.png"
    f2 = folder / "new.png"
    _make_png(f1)
    _make_png(f2)
    os.utime(f1, (1_700_000_000, 1_700_000_000))
    os.utime(f2, (1_800_000_000, 1_800_000_000))
    fake = FakeAutoProvider(detections=["ui"], analyses=["新图分析"])
    monkeypatch.setattr(tools_module, "get_provider", lambda s: fake)
    settings = Settings(
        auto_tile=False, max_pixels=1_000_000, cache_dir=tmp_path / "cache"
    )
    mcp = FastMCP("vision-mcp-test")
    tools_module.register_tools(mcp, settings)

    since = "2027-01-01T00:00:00"
    text = _call(mcp, "scan_folder", {"folder": str(folder), "since": since})
    assert "new.png" in text
    assert "old.png" not in text


class FakeMinerUClient:
    def __init__(self, settings):
        self.settings = settings

    async def parse(self, source, **kwargs):
        return {
            "state": "done",
            "source": source,
            "task_id": "t1",
            "batch_id": "",
            "markdown": "# 解析结果",
            "md_path": "C:/out/full.md",
        }

    async def status(self, task_id="", batch_id=""):
        return {
            "task_id": task_id,
            "batch_id": batch_id,
            "state": "done",
            "full_zip_url": "https://z/full.zip",
        }


def test_parse_document_tool_returns_markdown(monkeypatch, tmp_path):
    import tools as tools_module

    settings = Settings(
        mineru_api_key="sk-mineru", mineru_max_wait_s=30, output_dir=tmp_path
    )
    monkeypatch.setattr(tools_module, "MinerUClient", FakeMinerUClient)
    mcp = FastMCP("vision-mcp-test")
    tools_module.register_tools(mcp, settings)
    text = _call(mcp, "parse_document", {"source": "https://example.com/a.pdf"})
    assert "MinerU 解析完成" in text
    assert "# 解析结果" in text
    assert "C:/out/full.md" in text


def test_parse_document_status_tool(monkeypatch):
    import tools as tools_module

    settings = Settings(mineru_api_key="sk-mineru")
    monkeypatch.setattr(tools_module, "MinerUClient", FakeMinerUClient)
    mcp = FastMCP("vision-mcp-test")
    tools_module.register_tools(mcp, settings)
    text = _call(mcp, "parse_document_status", {"task_id": "t1"})
    assert "t1" in text
    assert "done" in text
